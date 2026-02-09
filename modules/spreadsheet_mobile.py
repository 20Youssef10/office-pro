"""
Spreadsheet Mobile Module
Kivy-based mobile-optimized spreadsheet
"""

import os
import csv
from pathlib import Path

try:
    from kivy.uix.screenmanager import Screen
    from kivy.uix.boxlayout import BoxLayout
    from kivy.uix.textinput import TextInput
    from kivy.uix.button import Button
    from kivy.uix.gridlayout import GridLayout
    from kivy.uix.scrollview import ScrollView
    from kivy.uix.popup import Popup
    from kivy.uix.label import Label
    from kivy.uix.filechooser import FileChooserListView
    from kivy.utils import platform

    KIVY_AVAILABLE = True
except ImportError:
    KIVY_AVAILABLE = False

try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class SpreadsheetMobile(Screen):
    """Mobile spreadsheet screen"""

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.current_file = None
        self.is_modified = False
        self.cells = {}  # Dictionary to store cell data
        self.current_row = 0
        self.current_col = 0
        self.build_ui()

    def build_ui(self):
        """Build mobile UI"""
        layout = BoxLayout(orientation="vertical")

        # Toolbar
        toolbar = GridLayout(cols=5, size_hint_y=0.08, spacing=2, padding=2)
        toolbar.add_widget(Button(text="New", on_press=self.new_sheet))
        toolbar.add_widget(Button(text="Open", on_press=self.open_file))
        toolbar.add_widget(Button(text="Save", on_press=self.save_file))
        toolbar.add_widget(Button(text="Clear", on_press=self.clear_cell))
        toolbar.add_widget(Button(text="Back", on_press=self.go_back))
        layout.add_widget(toolbar)

        # Cell reference and formula bar
        formula_layout = BoxLayout(size_hint_y=0.06)
        self.cell_ref_label = Label(text="A1", size_hint_x=0.15, font_size="14sp")
        formula_layout.add_widget(self.cell_ref_label)

        self.formula_input = TextInput(
            multiline=False, size_hint_x=0.85, font_size="14sp"
        )
        self.formula_input.bind(on_text_validate=self.apply_formula)
        formula_layout.add_widget(self.formula_input)
        layout.add_widget(formula_layout)

        # Spreadsheet grid
        scroll = ScrollView()
        self.grid = GridLayout(cols=5, spacing=2, padding=2, size_hint_y=None)
        self.grid.bind(minimum_height=self.grid.setter("height"))

        # Create cells
        self.create_grid()

        scroll.add_widget(self.grid)
        layout.add_widget(scroll)

        # Navigation buttons
        nav_layout = BoxLayout(size_hint_y=0.08)
        nav_layout.add_widget(Button(text="←", on_press=self.prev_cell))
        nav_layout.add_widget(Button(text="↑", on_press=self.up_cell))
        nav_layout.add_widget(Button(text="↓", on_press=self.down_cell))
        nav_layout.add_widget(Button(text="→", on_press=self.next_cell))
        layout.add_widget(nav_layout)

        self.add_widget(layout)

    def create_grid(self):
        """Create spreadsheet grid"""
        # Header row
        self.grid.add_widget(Label(text="", size_hint_y=None, height=40))
        for col in range(4):
            self.grid.add_widget(
                Label(
                    text=chr(65 + col),  # A, B, C, D
                    size_hint_y=None,
                    height=40,
                    bold=True,
                )
            )

        # Data rows
        for row in range(20):
            # Row header
            self.grid.add_widget(
                Label(text=str(row + 1), size_hint_y=None, height=50, bold=True)
            )

            # Cells
            for col in range(4):
                cell = TextInput(
                    multiline=False, size_hint_y=None, height=50, font_size="14sp"
                )
                cell.bind(focus=self.on_cell_focus)
                cell.row = row
                cell.col = col
                self.cells[(row, col)] = cell
                self.grid.add_widget(cell)

    def on_cell_focus(self, instance, value):
        """Handle cell focus"""
        if value:  # Cell got focus
            self.current_row = instance.row
            self.current_col = instance.col
            self.update_cell_ref()
            self.formula_input.text = instance.text

    def update_cell_ref(self):
        """Update cell reference label"""
        col_letter = chr(65 + self.current_col)
        self.cell_ref_label.text = f"{col_letter}{self.current_row + 1}"

    def apply_formula(self, instance):
        """Apply formula to current cell"""
        key = (self.current_row, self.current_col)
        if key in self.cells:
            self.cells[key].text = self.formula_input.text
            self.is_modified = True

    def next_cell(self, instance):
        """Move to next cell"""
        key = (self.current_row, self.current_col)
        if key in self.cells:
            self.cells[key].focus = False
        self.current_col = (self.current_col + 1) % 4
        key = (self.current_row, self.current_col)
        if key in self.cells:
            self.cells[key].focus = True

    def prev_cell(self, instance):
        """Move to previous cell"""
        key = (self.current_row, self.current_col)
        if key in self.cells:
            self.cells[key].focus = False
        self.current_col = (self.current_col - 1) % 4
        key = (self.current_row, self.current_col)
        if key in self.cells:
            self.cells[key].focus = True

    def up_cell(self, instance):
        """Move to cell above"""
        key = (self.current_row, self.current_col)
        if key in self.cells:
            self.cells[key].focus = False
        self.current_row = max(0, self.current_row - 1)
        key = (self.current_row, self.current_col)
        if key in self.cells:
            self.cells[key].focus = True

    def down_cell(self, instance):
        """Move to cell below"""
        key = (self.current_row, self.current_col)
        if key in self.cells:
            self.cells[key].focus = False
        self.current_row = min(19, self.current_row + 1)
        key = (self.current_row, self.current_col)
        if key in self.cells:
            self.cells[key].focus = True

    def clear_cell(self, instance):
        """Clear current cell"""
        key = (self.current_row, self.current_col)
        if key in self.cells:
            self.cells[key].text = ""
            self.formula_input.text = ""
            self.is_modified = True

    def new_sheet(self, instance):
        """Create new spreadsheet"""
        if self.is_modified:
            self.show_confirm_dialog()
        else:
            self.clear_all_cells()

    def clear_all_cells(self):
        """Clear all cells"""
        for cell in self.cells.values():
            cell.text = ""
        self.current_file = None
        self.is_modified = False

    def open_file(self, instance):
        """Open file dialog"""
        content = BoxLayout(orientation="vertical")
        filechooser = FileChooserListView(
            path=self.get_documents_path(), filters=["*.xlsx", "*.csv", "*.xls"]
        )
        content.add_widget(filechooser)

        btn_layout = BoxLayout(size_hint_y=0.1)
        btn_layout.add_widget(
            Button(
                text="Open", on_press=lambda x: self.load_file(filechooser.selection)
            )
        )
        btn_layout.add_widget(
            Button(text="Cancel", on_press=lambda x: self.popup.dismiss())
        )
        content.add_widget(btn_layout)

        self.popup = Popup(title="Open File", content=content, size_hint=(0.9, 0.9))
        self.popup.open()

    def get_documents_path(self):
        """Get documents path"""
        if platform == "android":
            from android.storage import primary_external_storage_path

            return primary_external_storage_path()
        else:
            return os.path.expanduser("~")

    def load_file(self, selection):
        """Load selected file"""
        if selection:
            file_path = selection[0]
            try:
                self.clear_all_cells()

                if file_path.endswith(".xlsx") and OPENPYXL_AVAILABLE:
                    self.load_xlsx(file_path)
                elif file_path.endswith(".csv"):
                    self.load_csv(file_path)

                self.current_file = file_path
                self.is_modified = False
                self.popup.dismiss()
                self.show_message("File loaded!")

            except Exception as e:
                self.show_error(f"Error loading file: {str(e)}")

    def load_xlsx(self, file_path):
        """Load Excel file"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        for row in range(min(20, ws.max_row)):
            for col in range(min(4, ws.max_column)):
                cell = ws.cell(row=row + 1, column=col + 1)
                if cell.value:
                    key = (row, col)
                    if key in self.cells:
                        self.cells[key].text = str(cell.value)

    def load_csv(self, file_path):
        """Load CSV file"""
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader):
                if row_idx >= 20:
                    break
                for col_idx, value in enumerate(row[:4]):
                    key = (row_idx, col_idx)
                    if key in self.cells:
                        self.cells[key].text = value

    def save_file(self, instance):
        """Save file"""
        if self.current_file:
            self.save_to_file(self.current_file)
        else:
            self.save_as_dialog()

    def save_as_dialog(self):
        """Save as dialog"""
        content = BoxLayout(orientation="vertical")

        filename_input = TextInput(
            hint_text="Enter filename", multiline=False, size_hint_y=0.1
        )
        content.add_widget(filename_input)

        btn_layout = BoxLayout(size_hint_y=0.1)
        btn_layout.add_widget(
            Button(
                text="Save",
                on_press=lambda x: self.perform_save_as(filename_input.text),
            )
        )
        btn_layout.add_widget(
            Button(text="Cancel", on_press=lambda x: self.popup.dismiss())
        )
        content.add_widget(btn_layout)

        self.popup = Popup(title="Save As", content=content, size_hint=(0.9, 0.5))
        self.popup.open()

    def perform_save_as(self, filename):
        """Perform save as"""
        if filename:
            if not filename.endswith((".xlsx", ".csv")):
                filename += ".xlsx"
            file_path = os.path.join(self.get_documents_path(), filename)
            self.save_to_file(file_path)
            self.popup.dismiss()

    def save_to_file(self, file_path):
        """Save to file"""
        try:
            if file_path.endswith(".xlsx") and OPENPYXL_AVAILABLE:
                self.save_xlsx(file_path)
            else:
                self.save_csv(file_path.replace(".xlsx", ".csv"))

            self.current_file = file_path
            self.is_modified = False
            self.show_message("Saved successfully!")

        except Exception as e:
            self.show_error(f"Error saving file: {str(e)}")

    def save_xlsx(self, file_path):
        """Save as Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active

        for (row, col), cell in self.cells.items():
            if cell.text:
                ws.cell(row=row + 1, column=col + 1, value=cell.text)

        wb.save(file_path)

    def save_csv(self, file_path):
        """Save as CSV"""
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in range(20):
                row_data = []
                for col in range(4):
                    key = (row, col)
                    if key in self.cells:
                        row_data.append(self.cells[key].text)
                    else:
                        row_data.append("")
                writer.writerow(row_data)

    def show_confirm_dialog(self):
        """Show confirmation dialog"""
        content = BoxLayout(orientation="vertical")
        content.add_widget(Label(text="Save changes?"))

        btn_layout = BoxLayout(size_hint_y=0.3)
        btn_layout.add_widget(
            Button(text="Save", on_press=lambda x: self.save_and_new())
        )
        btn_layout.add_widget(
            Button(text="Discard", on_press=lambda x: self.discard_and_new())
        )
        btn_layout.add_widget(
            Button(text="Cancel", on_press=lambda x: self.popup.dismiss())
        )
        content.add_widget(btn_layout)

        self.popup = Popup(
            title="Unsaved Changes", content=content, size_hint=(0.8, 0.4)
        )
        self.popup.open()

    def save_and_new(self):
        """Save and create new"""
        self.save_file(None)
        self.popup.dismiss()
        self.clear_all_cells()

    def discard_and_new(self):
        """Discard and create new"""
        self.popup.dismiss()
        self.clear_all_cells()

    def show_message(self, message):
        """Show message"""
        popup = Popup(
            title="Message", content=Label(text=message), size_hint=(0.7, 0.3)
        )
        popup.open()

    def show_error(self, error):
        """Show error"""
        popup = Popup(title="Error", content=Label(text=error), size_hint=(0.8, 0.4))
        popup.open()

    def go_back(self, instance):
        """Go back to main menu"""
        self.manager.current = "menu"

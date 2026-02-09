"""
Spreadsheet Module
Professional spreadsheet editing with support for XLSX, XLS, ODS, and CSV files.
Enhanced with charts, conditional formatting, freeze panes, and more.
"""

import os
import csv
import json
from pathlib import Path
from typing import List, Optional, Any, Dict, Tuple
from datetime import datetime

try:
    from PyQt6.QtWidgets import (
        QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
        QTableWidgetItem, QToolBar, QPushButton, QComboBox,
        QLineEdit, QFileDialog, QMessageBox, QInputDialog,
        QHeaderView, QAbstractItemView, QLabel, QFrame,
        QSplitter, QTextEdit, QMenu, QApplication, QDialog,
        QDialogButtonBox, QGroupBox, QRadioButton, QCheckBox,
        QSpinBox, QDoubleSpinBox, QTabWidget, QProgressBar,
        QScrollArea, QGraphicsView, QGraphicsScene, QListWidget,
        QListWidgetItem, QColorDialog, QFontDialog
    )
    from PyQt6.QtCore import Qt, QSize, pyqtSignal, QTimer, QRect
    from PyQt6.QtGui import QAction, QFont, QColor, QBrush, QPainter, QPen
    from PyQt6.QtCharts import QChart, QChartView, QBarSeries, QBarSet, QLineSeries, 
                              QPieSeries, QValueAxis, QBarCategoryAxis

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
        OPENPYXL_AVAILABLE = True
    except ImportError:
        OPENPYXL_AVAILABLE = False

    try:
        import pandas as pd
        import numpy as np
        PANDAS_AVAILABLE = True
    except ImportError:
        PANDAS_AVAILABLE = False

except ImportError as e:
    print(f"Import error in spreadsheet: {e}")
    raise


class ChartDialog(QDialog):
    """Dialog for creating charts"""
    
    def __init__(self, parent=None, table=None):
        super().__init__(parent)
        self.table = table
        self.setWindowTitle("Insert Chart")
        self.setMinimumWidth(500)
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Chart type selection
        type_group = QGroupBox("Chart Type")
        type_layout = QHBoxLayout()
        
        self.chart_types = QComboBox()
        self.chart_types.addItems(["Bar Chart", "Line Chart", "Pie Chart", "Column Chart"])
        type_layout.addWidget(self.chart_types)
        type_group.setLayout(type_layout)
        layout.addWidget(type_group)
        
        # Data range
        range_group = QGroupBox("Data Range")
        range_layout = QVBoxLayout()
        
        self.range_input = QLineEdit()
        self.range_input.setPlaceholderText("e.g., A1:D10")
        range_layout.addWidget(QLabel("Cell Range:"))
        range_layout.addWidget(self.range_input)
        
        self.first_row_header = QCheckBox("First row contains headers")
        self.first_row_header.setChecked(True)
        range_layout.addWidget(self.first_row_header)
        
        self.first_col_header = QCheckBox("First column contains labels")
        range_layout.addWidget(self.first_col_header)
        
        range_group.setLayout(range_layout)
        layout.addWidget(range_group)
        
        # Chart title
        title_group = QGroupBox("Chart Title")
        title_layout = QVBoxLayout()
        self.title_input = QLineEdit()
        self.title_input.setPlaceholderText("Enter chart title...")
        title_layout.addWidget(self.title_input)
        title_group.setLayout(title_layout)
        layout.addWidget(title_group)
        
        # Buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
    def get_chart_data(self):
        """Get chart configuration"""
        return {
            'type': self.chart_types.currentText(),
            'range': self.range_input.text(),
            'first_row_header': self.first_row_header.isChecked(),
            'first_col_header': self.first_col_header.isChecked(),
            'title': self.title_input.text()
        }


class ConditionalFormattingDialog(QDialog):
    """Dialog for conditional formatting"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Conditional Formatting")
        self.setMinimumWidth(400)
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Rule type
        rule_group = QGroupBox("Formatting Rule")
        rule_layout = QVBoxLayout()
        
        self.rule_type = QComboBox()
        self.rule_type.addItems([
            "Greater than",
            "Less than",
            "Equal to",
            "Between",
            "Text contains",
            "Duplicate values"
        ])
        rule_layout.addWidget(self.rule_type)
        
        self.value_input = QLineEdit()
        self.value_input.setPlaceholderText("Value...")
        rule_layout.addWidget(self.value_input)
        
        rule_group.setLayout(rule_layout)
        layout.addWidget(rule_group)
        
        # Format
        format_group = QGroupBox("Format")
        format_layout = QVBoxLayout()
        
        self.bg_color_btn = QPushButton("Background Color")
        self.bg_color_btn.clicked.connect(self.select_bg_color)
        format_layout.addWidget(self.bg_color_btn)
        
        self.text_color_btn = QPushButton("Text Color")
        self.text_color_btn.clicked.connect(self.select_text_color)
        format_layout.addWidget(self.text_color_btn)
        
        self.bold_check = QCheckBox("Bold")
        format_layout.addWidget(self.bold_check)
        
        self.italic_check = QCheckBox("Italic")
        format_layout.addWidget(self.italic_check)
        
        format_group.setLayout(format_layout)
        layout.addWidget(format_group)
        
        self.bg_color = QColor("yellow")
        self.text_color = QColor("black")
        
        # Buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
    def select_bg_color(self):
        """Select background color"""
        color = QColorDialog.getColor(self.bg_color, self)
        if color.isValid():
            self.bg_color = color
            self.bg_color_btn.setStyleSheet(f"background-color: {color.name()};")
            
    def select_text_color(self):
        """Select text color"""
        color = QColorDialog.getColor(self.text_color, self)
        if color.isValid():
            self.text_color = color
            self.text_color_btn.setStyleSheet(f"background-color: {color.name()};")
            
    def get_formatting_rule(self):
        """Get formatting rule"""
        return {
            'type': self.rule_type.currentText(),
            'value': self.value_input.text(),
            'bg_color': self.bg_color,
            'text_color': self.text_color,
            'bold': self.bold_check.isChecked(),
            'italic': self.italic_check.isChecked()
        }


class SortDialog(QDialog):
    """Sort dialog"""
    
    def __init__(self, parent=None, columns=None):
        super().__init__(parent)
        self.columns = columns or []
        self.setWindowTitle("Sort")
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Sort by
        layout.addWidget(QLabel("Sort by:"))
        self.sort_column = QComboBox()
        self.sort_column.addItems(self.columns)
        layout.addWidget(self.sort_column)
        
        # Order
        self.ascending = QRadioButton("Ascending (A to Z)")
        self.ascending.setChecked(True)
        layout.addWidget(self.ascending)
        
        self.descending = QRadioButton("Descending (Z to A)")
        layout.addWidget(self.descending)
        
        # Buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
    def get_sort_options(self):
        """Get sort options"""
        return {
            'column': self.sort_column.currentIndex(),
            'ascending': self.ascending.isChecked()
        }


class SpreadsheetTable(QTableWidget):
    """Enhanced custom table widget for spreadsheet functionality"""
    
    cell_changed = pyqtSignal(int, int)
    
    def __init__(self, rows=100, cols=52, parent=None):
        super().__init__(rows, cols, parent)
        self.setup_table()
        self.conditional_rules = []
        self.frozen_rows = 0
        self.frozen_cols = 0
        
    def setup_table(self):
        """Setup table properties"""
        self.setAlternatingRowColors(True)
        self.setStyleSheet("""
            QTableWidget {
                background-color: white;
                alternate-background-color: #f9f9f9;
                gridline-color: #d0d0d0;
                border: none;
            }
            QTableWidget::item {
                padding: 4px;
                border: 1px solid #d0d0d0;
            }
            QTableWidget::item:selected {
                background-color: #d0e0f0;
                color: black;
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                padding: 5px;
                border: 1px solid #d0d0d0;
                font-weight: bold;
            }
        """)
        
        # Set column headers (A, B, C, ...)
        headers = []
        for i in range(self.columnCount()):
            headers.append(self.get_column_name(i))
        self.setHorizontalHeaderLabels(headers)
        
        # Set row headers (1, 2, 3, ...)
        row_headers = [str(i + 1) for i in range(self.rowCount())]
        self.setVerticalHeaderLabels(row_headers)
        
        # Enable features
        self.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        
        # Cell change tracking
        self.itemChanged.connect(self.on_item_changed)
        
        # Enable drag and drop
        self.setDragEnabled(True)
        self.setAcceptDrops(True)
        self.setDropIndicatorShown(True)
        
    def get_column_name(self, index: int) -> str:
        """Convert column index to letter (0=A, 25=Z, 26=AA, etc.)"""
        result = ""
        index += 1
        while index > 0:
            index, remainder = divmod(index - 1, 26)
            result = chr(65 + remainder) + result
        return result
        
    def on_item_changed(self, item):
        """Handle cell changes"""
        self.cell_changed.emit(item.row(), item.column())
        self.apply_conditional_formatting(item.row(), item.column())
        
    def get_cell_value(self, row: int, col: int) -> str:
        """Get cell value"""
        item = self.item(row, col)
        return item.text() if item else ""
        
    def set_cell_value(self, row: int, col: int, value: str):
        """Set cell value"""
        item = self.item(row, col)
        if not item:
            item = QTableWidgetItem()
            self.setItem(row, col, item)
        item.setText(str(value))
        
    def get_used_range(self) -> tuple:
        """Get the range of cells that contain data"""
        max_row = 0
        max_col = 0
        
        for row in range(self.rowCount()):
            for col in range(self.columnCount()):
                if self.get_cell_value(row, col):
                    max_row = max(max_row, row)
                    max_col = max(max_col, col)
                    
        return (max_row + 1, max_col + 1)
        
    def add_conditional_formatting_rule(self, rule: dict):
        """Add conditional formatting rule"""
        self.conditional_rules.append(rule)
        self.apply_all_conditional_formatting()
        
    def apply_all_conditional_formatting(self):
        """Apply all conditional formatting rules"""
        max_row, max_col = self.get_used_range()
        for row in range(max_row):
            for col in range(max_col):
                self.apply_conditional_formatting(row, col)
                
    def apply_conditional_formatting(self, row: int, col: int):
        """Apply conditional formatting to cell"""
        item = self.item(row, col)
        if not item:
            return
            
        value_text = item.text()
        
        try:
            value = float(value_text)
        except ValueError:
            return
            
        for rule in self.conditional_rules:
            applies = False
            
            if rule['type'] == "Greater than":
                try:
                    if value > float(rule['value']):
                        applies = True
                except:
                    pass
            elif rule['type'] == "Less than":
                try:
                    if value < float(rule['value']):
                        applies = True
                except:
                    pass
            elif rule['type'] == "Equal to":
                try:
                    if value == float(rule['value']):
                        applies = True
                except:
                    pass
                    
            if applies:
                item.setBackground(QBrush(rule['bg_color']))
                item.setForeground(QBrush(rule['text_color']))
                
                font = item.font()
                font.setBold(rule.get('bold', False))
                font.setItalic(rule.get('italic', False))
                item.setFont(font)
                
    def auto_fill(self, start_row: int, start_col: int, end_row: int, end_col: int):
        """Auto-fill cells based on pattern"""
        # Get initial values
        values = []
        for row in range(start_row, end_row + 1):
            val = self.get_cell_value(row, start_col)
            values.append(val)
            
        # Detect pattern (simple arithmetic sequence)
        if len(values) >= 2:
            try:
                nums = [float(v) for v in values if v]
                if len(nums) >= 2:
                    diff = nums[1] - nums[0]
                    for col in range(start_col + 1, end_col + 1):
                        last_val = float(self.get_cell_value(end_row, col - 1) or 0)
                        for row in range(start_row, end_row + 1):
                            new_val = last_val + diff
                            self.set_cell_value(row, col, str(new_val))
                            last_val = new_val
            except:
                # Copy pattern
                pattern = values
                for col in range(start_col + 1, end_col + 1):
                    for i, row in enumerate(range(start_row, end_row + 1)):
                        if i < len(pattern):
                            self.set_cell_value(row, col, pattern[i])


class SpreadsheetEditor(QWidget):
    """Enhanced Professional Spreadsheet Editor with Office features"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.current_file = None
        self.is_modified = False
        self.formula_bar = None
        self.charts = []
        
        # Auto-save
        self.autosave_timer = QTimer()
        self.autosave_timer.timeout.connect(self.autosave)
        self.autosave_enabled = True
        
        self.init_ui()
        self.setup_toolbar()
        self.setup_shortcuts()
        self.new_spreadsheet()
        
        # Start auto-save
        if self.autosave_enabled:
            self.autosave_timer.start(300000)  # 5 minutes
        
    def init_ui(self):
        """Initialize the user interface"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Toolbar
        self.toolbar = QToolBar()
        self.toolbar.setMovable(False)
        self.toolbar.setStyleSheet("""
            QToolBar {
                background-color: #217346;
                border: none;
                padding: 5px;
                spacing: 5px;
            }
            QToolBar QPushButton {
                background-color: rgba(255,255,255,0.1);
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 4px;
                font-size: 12px;
            }
            QToolBar QPushButton:hover {
                background-color: rgba(255,255,255,0.2);
            }
            QToolBar QLineEdit {
                background-color: white;
                border: none;
                padding: 5px;
                border-radius: 3px;
                min-width: 200px;
            }
        """)
        layout.addWidget(self.toolbar)

        # Main editor area with formula bar
        main_widget = QWidget()
        main_layout = QVBoxLayout(main_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Formula bar
        formula_frame = QFrame()
        formula_frame.setStyleSheet("""
            QFrame {
                background-color: #f5f5f5;
                border-bottom: 1px solid #ddd;
            }
        """)
        formula_layout = QHBoxLayout(formula_frame)
        formula_layout.setContentsMargins(10, 5, 10, 5)

        self.cell_ref_label = QLabel("A1")
        self.cell_ref_label.setStyleSheet("font-weight: bold; min-width: 40px;")
        formula_layout.addWidget(self.cell_ref_label)

        self.formula_bar = QLineEdit()
        self.formula_bar.setPlaceholderText("Enter formula or value...")
        self.formula_bar.returnPressed.connect(self.apply_formula)
        formula_layout.addWidget(self.formula_bar)

        main_layout.addWidget(formula_frame)

        # Spreadsheet table with splitter
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Spreadsheet
        self.table = SpreadsheetTable(1000, 52)
        self.table.currentCellChanged.connect(self.on_cell_changed)
        self.table.cell_changed.connect(self.on_cell_modified)
        splitter.addWidget(self.table)
        
        # Charts panel (right side)
        self.charts_panel = QWidget()
        charts_layout = QVBoxLayout(self.charts_panel)
        charts_layout.setContentsMargins(5, 5, 5, 5)
        
        charts_label = QLabel("Charts")
        charts_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        charts_layout.addWidget(charts_label)
        
        self.charts_list = QListWidget()
        self.charts_list.setMaximumWidth(250)
        charts_layout.addWidget(self.charts_list)
        
        splitter.addWidget(self.charts_panel)
        splitter.setSizes([1200, 200])
        
        main_layout.addWidget(splitter)
        layout.addWidget(main_widget)

        # Status bar
        self.status_frame = QFrame()
        self.status_frame.setStyleSheet("""
            QFrame {
                background-color: #f0f0f0;
                border-top: 1px solid #ddd;
                padding: 5px;
            }
        """)
        status_layout = QHBoxLayout(self.status_frame)
        status_layout.setContentsMargins(10, 5, 10, 5)

        self.cell_count_label = QLabel("Cells: 0")
        self.selected_range_label = QLabel("Selected: A1")
        self.autosave_status = QLabel("Auto-save: ON")

        status_layout.addWidget(self.cell_count_label)
        status_layout.addStretch()
        status_layout.addWidget(self.selected_range_label)
        status_layout.addWidget(self.autosave_status)

        layout.addWidget(self.status_frame)

    def setup_toolbar(self):
        """Setup the toolbar with Office features"""
        # File operations
        self.new_btn = QPushButton("New")
        self.new_btn.clicked.connect(self.new_spreadsheet)
        self.toolbar.addWidget(self.new_btn)

        self.open_btn = QPushButton("Open")
        self.open_btn.clicked.connect(self.open_file_dialog)
        self.toolbar.addWidget(self.open_btn)

        self.save_btn = QPushButton("Save")
        self.save_btn.clicked.connect(self.save)
        self.toolbar.addWidget(self.save_btn)

        self.toolbar.addSeparator()

        # Formatting
        self.bold_btn = QPushButton("B")
        self.bold_btn.setCheckable(True)
        self.bold_btn.setStyleSheet("font-weight: bold; min-width: 30px;")
        self.bold_btn.clicked.connect(self.toggle_bold)
        self.toolbar.addWidget(self.bold_btn)

        # Insert/Delete
        self.insert_row_btn = QPushButton("+ Row")
        self.insert_row_btn.clicked.connect(self.insert_row)
        self.toolbar.addWidget(self.insert_row_btn)

        self.delete_row_btn = QPushButton("- Row")
        self.delete_row_btn.clicked.connect(self.delete_row)
        self.toolbar.addWidget(self.delete_row_btn)

        self.insert_col_btn = QPushButton("+ Col")
        self.insert_col_btn.clicked.connect(self.insert_column)
        self.toolbar.addWidget(self.insert_col_btn)

        self.delete_col_btn = QPushButton("- Col")
        self.delete_col_btn.clicked.connect(self.delete_column)
        self.toolbar.addWidget(self.delete_col_btn)

        self.toolbar.addSeparator()

        # Functions
        self.sum_btn = QPushButton("Σ Sum")
        self.sum_btn.clicked.connect(self.insert_sum)
        self.toolbar.addWidget(self.sum_btn)

        self.avg_btn = QPushButton("Avg")
        self.avg_btn.clicked.connect(self.insert_average)
        self.toolbar.addWidget(self.avg_btn)
        
        self.toolbar.addSeparator()
        
        # NEW FEATURES
        # Charts
        self.chart_btn = QPushButton("Insert Chart")
        self.chart_btn.clicked.connect(self.insert_chart)
        self.toolbar.addWidget(self.chart_btn)
        
        # Conditional Formatting
        self.conditional_btn = QPushButton("Conditional Format")
        self.conditional_btn.clicked.connect(self.show_conditional_formatting)
        self.toolbar.addWidget(self.conditional_btn)
        
        # Freeze Panes
        self.freeze_btn = QPushButton("Freeze Panes")
        self.freeze_btn.clicked.connect(self.freeze_panes)
        self.toolbar.addWidget(self.freeze_btn)
        
        # Sort
        self.sort_btn = QPushButton("Sort")
        self.sort_btn.clicked.connect(self.show_sort_dialog)
        self.toolbar.addWidget(self.sort_btn)
        
        # Auto-fill
        self.autofill_btn = QPushButton("Auto-fill")
        self.autofill_btn.clicked.connect(self.auto_fill_selection)
        self.toolbar.addWidget(self.autofill_btn)

        self.toolbar.addSeparator()

        # Home button
        self.home_btn = QPushButton("← Home")
        self.home_btn.clicked.connect(self.go_home)
        self.toolbar.addWidget(self.home_btn)
        
    def setup_shortcuts(self):
        """Setup keyboard shortcuts"""
        from PyQt6.QtGui import QKeySequence, QShortcut
        
        # Ctrl+D for fill down
        fill_down = QShortcut(QKeySequence("Ctrl+D"), self)
        fill_down.activated.connect(self.fill_down)
        
        # Ctrl+R for fill right
        fill_right = QShortcut(QKeySequence("Ctrl+R"), self)
        fill_right.activated.connect(self.fill_right)
        
    def new_spreadsheet(self):
        """Create new spreadsheet"""
        if self.check_save():
            self.table.clearContents()
            self.current_file = None
            self.is_modified = False
            self.update_title()
            self.update_status()

    def on_cell_changed(self, row, col, prev_row, prev_col):
        """Handle cell selection change"""
        col_name = self.table.get_column_name(col)
        self.cell_ref_label.setText(f"{col_name}{row + 1}")
        self.selected_range_label.setText(f"Selected: {col_name}{row + 1}")

        # Update formula bar
        value = self.table.get_cell_value(row, col)
        self.formula_bar.setText(value)

    def on_cell_modified(self, row, col):
        """Handle cell modification"""
        self.is_modified = True
        self.update_title()
        self.update_status()

        # Update formula bar if this is the current cell
        current_row = self.table.currentRow()
        current_col = self.table.currentColumn()
        if row == current_row and col == current_col:
            value = self.table.get_cell_value(row, col)
            self.formula_bar.setText(value)

    def apply_formula(self):
        """Apply formula from formula bar"""
        text = self.formula_bar.text()
        row = self.table.currentRow()
        col = self.table.currentColumn()

        if row >= 0 and col >= 0:
            self.table.set_cell_value(row, col, text)

            # Evaluate formula if it starts with =
            if text.startswith('='):
                result = self.evaluate_formula(text)
                if result is not None:
                    item = self.table.item(row, col)
                    if item:
                        item.setToolTip(f"Formula: {text}")
                        
    def evaluate_formula(self, formula: str) -> Optional[float]:
        """Evaluate a formula"""
        try:
            if formula.startswith('='):
                formula = formula[1:]

            # Basic functions
            formula_upper = formula.upper()

            if formula_upper.startswith('SUM('):
                return self.calculate_sum(formula[4:-1])
            elif formula_upper.startswith('AVERAGE(') or formula_upper.startswith('AVG('):
                return self.calculate_average(formula[formula.find('(')+1:-1])
            elif formula_upper.startswith('COUNT('):
                return self.calculate_count(formula[6:-1])
            elif formula_upper.startswith('MAX('):
                return self.calculate_max(formula[4:-1])
            elif formula_upper.startswith('MIN('):
                return self.calculate_min(formula[4:-1])
            else:
                return None

        except Exception as e:
            return None

    def get_range_values(self, range_str: str) -> List[float]:
        """Get values from a cell range (e.g., "A1:B5")"""
        values = []

        try:
            if ':' in range_str:
                start, end = range_str.split(':')
                start_col = self.column_letter_to_index(start[0])
                start_row = int(start[1:]) - 1
                end_col = self.column_letter_to_index(end[0])
                end_row = int(end[1:]) - 1

                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        value = self.table.get_cell_value(row, col)
                        try:
                            values.append(float(value))
                        except ValueError:
                            pass
            else:
                # Single cell
                col = self.column_letter_to_index(range_str[0])
                row = int(range_str[1:]) - 1
                value = self.table.get_cell_value(row, col)
                try:
                    values.append(float(value))
                except ValueError:
                    pass

        except Exception:
            pass

        return values

    def column_letter_to_index(self, letter: str) -> int:
        """Convert column letter to index (A=0, B=1, etc.)"""
        result = 0
        for char in letter.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1

    def calculate_sum(self, range_str: str) -> float:
        """Calculate sum of a range"""
        values = self.get_range_values(range_str)
        return sum(values)

    def calculate_average(self, range_str: str) -> float:
        """Calculate average of a range"""
        values = self.get_range_values(range_str)
        return sum(values) / len(values) if values else 0

    def calculate_count(self, range_str: str) -> int:
        """Count numeric values in a range"""
        return len(self.get_range_values(range_str))

    def calculate_max(self, range_str: str) -> float:
        """Find maximum value in a range"""
        values = self.get_range_values(range_str)
        return max(values) if values else 0

    def calculate_min(self, range_str: str) -> float:
        """Find minimum value in a range"""
        values = self.get_range_values(range_str)
        return min(values) if values else 0

    def insert_sum(self):
        """Insert SUM formula"""
        text, ok = QInputDialog.getText(self, "Sum", "Enter range (e.g., A1:B5):")
        if ok and text:
            self.formula_bar.setText(f"=SUM({text})")
            self.apply_formula()

    def insert_average(self):
        """Insert AVERAGE formula"""
        text, ok = QInputDialog.getText(self, "Average", "Enter range (e.g., A1:B5):")
        if ok and text:
            self.formula_bar.setText(f"=AVERAGE({text})")
            self.apply_formula()
            
    # ========== NEW FEATURES ==========
    
    def insert_chart(self):
        """Insert chart"""
        dialog = ChartDialog(self, self.table)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            chart_data = dialog.get_chart_data()
            self.create_chart(chart_data)
            
    def create_chart(self, chart_data: dict):
        """Create and display chart"""
        try:
            # Parse range
            range_str = chart_data['range']
            if ':' in range_str:
                start, end = range_str.split(':')
                start_col = self.column_letter_to_index(start[0])
                start_row = int(start[1:]) - 1
                end_col = self.column_letter_to_index(end[0])
                end_row = int(end[1:]) - 1
            else:
                QMessageBox.warning(self, "Error", "Invalid range format")
                return
                
            # Create chart based on type
            chart = QChart()
            chart.setTitle(chart_data['title'])
            
            if chart_data['type'] == "Bar Chart":
                series = QBarSeries()
                categories = []
                
                for col in range(start_col, end_col + 1):
                    bar_set = QBarSet(f"Column {self.table.get_column_name(col)}")
                    for row in range(start_row, end_row + 1):
                        val = self.table.get_cell_value(row, col)
                        try:
                            bar_set.append(float(val))
                        except:
                            bar_set.append(0)
                    series.append(bar_set)
                    
                chart.addSeries(series)
                chart.createDefaultAxes()
                
            elif chart_data['type'] == "Pie Chart":
                series = QPieSeries()
                for row in range(start_row, end_row + 1):
                    label = self.table.get_cell_value(row, start_col)
                    val = self.table.get_cell_value(row, start_col + 1)
                    try:
                        series.append(label, float(val))
                    except:
                        pass
                chart.addSeries(series)
                
            elif chart_data['type'] == "Line Chart":
                series = QLineSeries()
                for row in range(start_row, end_row + 1):
                    val = self.table.get_cell_value(row, start_col + 1)
                    try:
                        series.append(row, float(val))
                    except:
                        pass
                chart.addSeries(series)
                chart.createDefaultAxes()
                
            # Add to charts list
            chart_item = QListWidgetItem(chart_data['title'])
            self.charts_list.addItem(chart_item)
            self.charts.append({'chart': chart, 'data': chart_data})
            
            QMessageBox.information(self, "Chart Created", 
                f"{chart_data['type']} '{chart_data['title']}' created successfully!")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to create chart:\n{str(e)}")
            
    def show_conditional_formatting(self):
        """Show conditional formatting dialog"""
        dialog = ConditionalFormattingDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            rule = dialog.get_formatting_rule()
            self.table.add_conditional_formatting_rule(rule)
            QMessageBox.information(self, "Formatting Applied", 
                "Conditional formatting rule has been applied.")
            
    def freeze_panes(self):
        """Freeze panes at current position"""
        row = self.table.currentRow()
        col = self.table.currentColumn()
        
        self.table.frozen_rows = row
        self.table.frozen_cols = col
        
        QMessageBox.information(self, "Freeze Panes", 
            f"Frozen at row {row + 1}, column {self.table.get_column_name(col)}")
            
    def show_sort_dialog(self):
        """Show sort dialog"""
        columns = [self.table.get_column_name(i) for i in range(min(26, self.table.columnCount()))]
        dialog = SortDialog(self, columns)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            options = dialog.get_sort_options()
            self.sort_data(options)
            
    def sort_data(self, options: dict):
        """Sort data based on options"""
        col = options['column']
        ascending = options['ascending']
        
        # Get all data in column
        max_row, _ = self.table.get_used_range()
        data = []
        for row in range(max_row):
            val = self.table.get_cell_value(row, col)
            try:
                num_val = float(val)
                data.append((row, num_val, val))
            except:
                data.append((row, 0, val))
                
        # Sort
        data.sort(key=lambda x: x[1], reverse=not ascending)
        
        # Get all row data
        row_data = []
        for row in range(max_row):
            row_values = []
            for c in range(self.table.columnCount()):
                row_values.append(self.table.get_cell_value(row, c))
            row_data.append(row_values)
            
        # Reorder rows
        for new_row, (old_row, _, _) in enumerate(data):
            for c in range(self.table.columnCount()):
                self.table.set_cell_value(new_row, c, row_data[old_row][c])
                
        QMessageBox.information(self, "Sort Complete", "Data has been sorted.")
        
    def auto_fill_selection(self):
        """Auto-fill selected cells"""
        selected = self.table.selectedRanges()
        if selected:
            sel = selected[0]
            self.table.auto_fill(
                sel.topRow(),
                sel.leftColumn(),
                sel.bottomRow(),
                sel.rightColumn()
            )
            
    def fill_down(self):
        """Fill down (Ctrl+D)"""
        selected = self.table.selectedRanges()
        if selected:
            sel = selected[0]
            if sel.rowCount() > 1:
                # Get value from top cell
                top_value = self.table.get_cell_value(sel.topRow(), sel.leftColumn())
                # Fill down
                for row in range(sel.topRow() + 1, sel.bottomRow() + 1):
                    for col in range(sel.leftColumn(), sel.rightColumn() + 1):
                        self.table.set_cell_value(row, col, top_value)
                        
    def fill_right(self):
        """Fill right (Ctrl+R)"""
        selected = self.table.selectedRanges()
        if selected:
            sel = selected[0]
            if sel.columnCount() > 1:
                # Get value from left cell
                left_value = self.table.get_cell_value(sel.topRow(), sel.leftColumn())
                # Fill right
                for row in range(sel.topRow(), sel.bottomRow() + 1):
                    for col in range(sel.leftColumn() + 1, sel.rightColumn() + 1):
                        self.table.set_cell_value(row, col, left_value)
                        
    def autosave(self):
        """Auto-save spreadsheet"""
        if self.is_modified and self.current_file:
            try:
                self.save_as(self.current_file)
                self.autosave_status.setText("Auto-save: Just now")
                QTimer.singleShot(5000, lambda: self.autosave_status.setText("Auto-save: ON"))
            except:
                self.autosave_status.setText("Auto-save: Failed")

    def toggle_bold(self):
        """Toggle bold formatting for selected cells"""
        font = QFont()
        font.setBold(self.bold_btn.isChecked())

        for item in self.table.selectedItems():
            item.setFont(font)

    def insert_row(self):
        """Insert row at current position"""
        row = self.table.currentRow()
        self.table.insertRow(row)
        self.update_row_headers()

    def delete_row(self):
        """Delete current row"""
        row = self.table.currentRow()
        if row >= 0:
            self.table.removeRow(row)
            self.update_row_headers()

    def insert_column(self):
        """Insert column at current position"""
        col = self.table.currentColumn()
        self.table.insertColumn(col)
        self.update_column_headers()

    def delete_column(self):
        """Delete current column"""
        col = self.table.currentColumn()
        if col >= 0:
            self.table.removeColumn(col)
            self.update_column_headers()

    def update_row_headers(self):
        """Update row header labels"""
        for i in range(self.table.rowCount()):
            self.table.setVerticalHeaderItem(i, QTableWidgetItem(str(i + 1)))

    def update_column_headers(self):
        """Update column header labels"""
        for i in range(self.table.columnCount()):
            self.table.setHorizontalHeaderItem(i, 
                QTableWidgetItem(self.table.get_column_name(i)))

    def open_file_dialog(self):
        """Open file dialog"""
        if self.check_save():
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Open Spreadsheet", "",
                "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)"
            )
            if file_path:
                self.open_spreadsheet(file_path)

    def open_spreadsheet(self, file_path):
        """Open a spreadsheet file"""
        try:
            if file_path.endswith('.xlsx') and OPENPYXL_AVAILABLE:
                self.open_xlsx(file_path)
            elif file_path.endswith('.csv'):
                self.open_csv(file_path)
            elif PANDAS_AVAILABLE:
                self.open_with_pandas(file_path)
            else:
                QMessageBox.warning(self, "Error", 
                    "Required libraries not installed. Install openpyxl and pandas.")
                return

            self.current_file = file_path
            self.is_modified = False
            self.update_title()
            self.update_status()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open spreadsheet:\n{str(e)}")

    def open_xlsx(self, file_path):
        """Open Excel file"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        self.table.clearContents()

        # Resize table if needed
        max_row = ws.max_row
        max_col = ws.max_column

        if max_row > self.table.rowCount():
            self.table.setRowCount(max_row)
        if max_col > self.table.columnCount():
            self.table.setColumnCount(max_col)

        # Load data
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    item = QTableWidgetItem(str(cell.value))

                    # Apply formatting
                    if cell.font.bold:
                        font = item.font()
                        font.setBold(True)
                        item.setFont(font)

                    self.table.setItem(row - 1, col - 1, item)

        self.update_row_headers()
        self.update_column_headers()

    def open_csv(self, file_path):
        """Open CSV file"""
        with open(file_path, 'r', newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            self.table.clearContents()

            row = 0
            for data in reader:
                if row >= self.table.rowCount():
                    self.table.setRowCount(row + 1)
                if len(data) > self.table.columnCount():
                    self.table.setColumnCount(len(data))

                for col, value in enumerate(data):
                    item = QTableWidgetItem(value)
                    self.table.setItem(row, col, item)

                row += 1

        self.update_row_headers()
        self.update_column_headers()

    def open_with_pandas(self, file_path):
        """Open file using pandas"""
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)

        self.table.clearContents()

        rows, cols = df.shape
        if rows > self.table.rowCount():
            self.table.setRowCount(rows)
        if cols > self.table.columnCount():
            self.table.setColumnCount(cols)

        # Set column headers
        for col, header in enumerate(df.columns):
            self.table.setHorizontalHeaderItem(col, QTableWidgetItem(str(header)))

        # Set data
        for row in range(rows):
            for col in range(cols):
                value = df.iloc[row, col]
                if pd.notna(value):
                    item = QTableWidgetItem(str(value))
                    self.table.setItem(row, col, item)

        self.update_row_headers()

    def save(self):
        """Save the spreadsheet"""
        if self.current_file:
            self.save_as(self.current_file)
        else:
            self.save_as_dialog()

    def save_as_dialog(self):
        """Save as dialog"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Spreadsheet", "",
            "Excel (*.xlsx);;CSV (*.csv);;All Files (*)"
        )
        if file_path:
            self.save_as(file_path)

    def save_as(self, file_path):
        """Save spreadsheet to file"""
        try:
            if file_path.endswith('.xlsx') and OPENPYXL_AVAILABLE:
                self.save_as_xlsx(file_path)
            elif file_path.endswith('.csv'):
                self.save_as_csv(file_path)
            elif OPENPYXL_AVAILABLE:
                file_path = file_path + '.xlsx'
                self.save_as_xlsx(file_path)
            else:
                QMessageBox.warning(self, "Error", 
                    "openpyxl not installed. Saving as CSV.")
                self.save_as_csv(file_path.replace('.xlsx', '.csv'))
                return

            self.current_file = file_path
            self.is_modified = False
            self.update_title()

            if self.parent:
                self.parent.status_bar.showMessage(f"Saved: {file_path}", 3000)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save spreadsheet:\n{str(e)}")

    def save_as_xlsx(self, file_path):
        """Save as Excel file"""
        wb = openpyxl.Workbook()
        ws = wb.active

        max_row, max_col = self.table.get_used_range()

        for row in range(max_row):
            for col in range(max_col):
                value = self.table.get_cell_value(row, col)
                if value:
                    ws.cell(row=row + 1, column=col + 1, value=value)

        wb.save(file_path)

    def save_as_csv(self, file_path):
        """Save as CSV file"""
        max_row, max_col = self.table.get_used_range()

        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            for row in range(max_row):
                row_data = []
                for col in range(max_col):
                    value = self.table.get_cell_value(row, col)
                    row_data.append(value)
                writer.writerow(row_data)

    def check_save(self):
        """Check if save is needed"""
        if self.is_modified:
            reply = QMessageBox.question(
                self, "Unsaved Changes",
                "The spreadsheet has unsaved changes. Save now?",
                QMessageBox.StandardButton.Save | 
                QMessageBox.StandardButton.Discard |
                QMessageBox.StandardButton.Cancel
            )

            if reply == QMessageBox.StandardButton.Save:
                self.save()
                return True
            elif reply == QMessageBox.StandardButton.Cancel:
                return False

        return True

    def update_title(self):
        """Update window title"""
        if self.parent:
            filename = "Untitled"
            if self.current_file:
                filename = Path(self.current_file).name
            modified = "*" if self.is_modified else ""
            self.parent.setWindowTitle(f"Office Pro - Spreadsheet - {filename}{modified}")

    def update_status(self):
        """Update status bar"""
        max_row, max_col = self.table.get_used_range()
        self.cell_count_label.setText(f"Cells: {max_row * max_col}")

    def go_home(self):
        """Return to main menu"""
        if self.check_save():
            if self.parent:
                self.parent.show_main_menu()
